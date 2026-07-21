from __future__ import annotations

import contextlib
import csv
import http.server
import io
import json
import re
import sys
import tempfile
import threading
import unittest
from pathlib import Path
from typing import Any, Iterator
from unittest import mock


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PACKAGE_ROOT))

import annotate_facebook_threads as pipeline  # noqa: E402


SCHEMA_PATH = PACKAGE_ROOT / "config" / "annotation_schema.json"
INPUT_XLSX = PACKAGE_ROOT / "input" / "facebook_comments_comprehensive_final.xlsx"
PROMPT_FILE = PACKAGE_ROOT / "input" / "ai_annotation_prompt.md"
START_SCRIPT = PACKAGE_ROOT / "scripts" / "start_vllm_8gpu_qwen32b.sh"
LEGACY_ERRORS_CSV = PACKAGE_ROOT / "output" / "qwen32b_8gpu" / "results" / "errors.csv"
LEGACY_ANNOTATIONS_CSV = (
    PACKAGE_ROOT / "output" / "qwen32b_8gpu" / "results" / "annotations.csv"
)


def make_row(
    excel_row: int = 2,
    *,
    topic: str = "carbon tax",
    conversation: str = "1. Alice [ROOT COMMENT]: I oppose the carbon tax.",
    message_count: int = 1,
) -> pipeline.InputRow:
    values: dict[str, Any] = {
        "topic": topic,
        "retrieval_keyword": topic,
        "post_text": "The post supports a carbon tax.",
        "post_url": f"https://example.test/post/{excel_row}",
        "post_stance_label": "support",
        "post_stance_target": topic,
        "post_mid": f"post-{excel_row}",
        "thread_index": excel_row - 1,
        "thread_id": f"thread-{excel_row}",
        "participants": "Alice",
        "commenter_ids": f"user-{excel_row}",
        "commenter_usernames": "Alice",
        "conversation_text": conversation,
        "root_comment_url": f"https://example.test/comment/{excel_row}",
        "message_count": message_count,
        "reply_count": max(0, message_count - 1),
    }
    records = tuple(
        {
            "data_type": "s" if isinstance(values[header], str) else "n",
            "value_type": type(values[header]).__name__,
            "value": values[header],
        }
        for header in pipeline.INPUT_HEADERS
    )
    input_hash = pipeline.hash_cell_records(records)
    business_key = f"{values['post_mid']}\x1f{values['thread_id']}"
    return pipeline.InputRow(
        excel_row=excel_row,
        values=values,
        cell_records=records,
        input_hash=input_hash,
        business_key=business_key,
        business_key_hash=pipeline.sha256_bytes(business_key.encode("utf-8")),
        task_key=pipeline.make_task_key(
            business_key,
            input_hash,
            "effective-prompt-hash",
            "schema-hash",
            "Qwen3-32B",
        ),
    )


def valid_related_annotation() -> dict[str, str]:
    return {
        "topic_relevance": "strongly_relevant",
        "training_grade": "generally_usable",
        "annotation_reason": "The comment directly discusses the carbon-tax topic and is interpretable.",
    }


def valid_unrelated_annotation() -> dict[str, str]:
    return {
        "topic_relevance": "off_topic",
        "training_grade": "unusable",
        "annotation_reason": "The comment has no explainable relationship to the assigned topic.",
    }


class _FakeVLLM:
    def __init__(self, planned: list[Any], model_id: str = "Qwen3-32B"):
        self.planned = list(planned)
        self.model_id = model_id
        self.requests: list[dict[str, Any]] = []
        self._lock = threading.Lock()
        outer = self

        class Handler(http.server.BaseHTTPRequestHandler):
            def log_message(self, _format: str, *args: Any) -> None:
                return

            def do_GET(self) -> None:  # noqa: N802
                if self.path.endswith("/models"):
                    body = json.dumps({"data": [{"id": outer.model_id}]}).encode()
                    self.send_response(200)
                    self.send_header("Content-Type", "application/json")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                    return
                self.send_error(404)

            def do_POST(self) -> None:  # noqa: N802
                length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(length).decode("utf-8"))
                with outer._lock:
                    request_index = len(outer.requests)
                    outer.requests.append(payload)
                    item = outer.planned[min(request_index, len(outer.planned) - 1)]

                if isinstance(item, tuple):
                    status, raw_body = item
                    body = raw_body.encode("utf-8")
                    self.send_response(status)
                    self.send_header("Content-Type", "application/json")
                    self.send_header("Content-Length", str(len(body)))
                    self.end_headers()
                    self.wfile.write(body)
                    return

                content = item if isinstance(item, str) else pipeline.canonical_json(item)
                response = {
                    "choices": [{"message": {"content": content}}],
                    "usage": {"prompt_tokens": 10, "completion_tokens": 5},
                }
                body = json.dumps(response, ensure_ascii=False).encode("utf-8")
                self.send_response(200)
                self.send_header("Content-Type", "application/json")
                self.send_header("Content-Length", str(len(body)))
                self.end_headers()
                self.wfile.write(body)

        self.server = http.server.ThreadingHTTPServer(("127.0.0.1", 0), Handler)
        self.thread = threading.Thread(target=self.server.serve_forever, daemon=True)
        self.thread.start()

    @property
    def base_url(self) -> str:
        host, port = self.server.server_address
        return f"http://{host}:{port}/v1"

    def close(self) -> None:
        self.server.shutdown()
        self.server.server_close()
        self.thread.join(timeout=5)


@contextlib.contextmanager
def fake_vllm(planned: list[Any], model_id: str = "Qwen3-32B") -> Iterator[_FakeVLLM]:
    server = _FakeVLLM(planned, model_id=model_id)
    try:
        yield server
    finally:
        server.close()


def make_client(server: _FakeVLLM, schema: dict[str, Any], *, network_retries: int = 0) -> pipeline.VLLMClient:
    return pipeline.VLLMClient(
        endpoint_pool=pipeline.EndpointPool([server.base_url]),
        model="Qwen3-32B",
        schema=schema,
        timeout=5,
        network_retries=network_retries,
        max_output_tokens=768,
        enable_thinking=False,
        api_key=None,
    )


class ValidationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.schema, _ = pipeline.load_schema(SCHEMA_PATH)

    def test_valid_annotation_passes(self) -> None:
        errors, warnings = pipeline.validate_annotation(
            valid_related_annotation(), make_row(), self.schema
        )
        self.assertEqual(errors, [])
        self.assertEqual(warnings, [])

    def test_enum_whitespace_is_rejected(self) -> None:
        value = valid_related_annotation()
        value["topic_relevance"] = "strongly_relevant "
        errors, _ = pipeline.validate_annotation(value, make_row(), self.schema)
        self.assertTrue(any("topic_relevance" in error for error in errors), errors)

    def test_relevance_grade_mismatch_is_rejected(self) -> None:
        value = valid_related_annotation()
        value["training_grade"] = "unusable"
        errors, _ = pipeline.validate_annotation(value, make_row(), self.schema)
        self.assertTrue(any("cross.relevant_grade" in error for error in errors), errors)

    def test_stance_fields_are_rejected_as_extra_fields(self) -> None:
        value = valid_related_annotation()
        value["stance_change"] = "no_evidence_of_change"
        errors, _ = pipeline.validate_annotation(value, make_row(), self.schema)
        self.assertTrue(any("schema.extra_fields" in error for error in errors), errors)

    def test_removed_field_is_rejected(self) -> None:
        value = valid_related_annotation()
        value["confidence"] = "high"
        errors, _ = pipeline.validate_annotation(value, make_row(), self.schema)
        self.assertTrue(any("schema.extra_fields" in error for error in errors), errors)

    def test_annotation_reason_over_240_characters_warns(self) -> None:
        value = valid_related_annotation()
        value["annotation_reason"] = "a" * 241
        errors, warnings = pipeline.validate_annotation(value, make_row(), self.schema)
        self.assertEqual(errors, [])
        self.assertTrue(any("reason.over_240" in warning for warning in warnings), warnings)

    def test_nontext_content_is_not_locally_reclassified(self) -> None:
        row = make_row(
            conversation="1. Alice [ROOT COMMENT]: [NON-TEXT COMMENT]",
            message_count=1,
        )
        errors, _ = pipeline.validate_annotation(valid_related_annotation(), row, self.schema)
        self.assertEqual(errors, [])


class IdentityAndSamplingTests(unittest.TestCase):
    def test_duration_formatting_covers_unknown_zero_and_hours(self) -> None:
        self.assertEqual(pipeline.format_duration(None), "unknown")
        self.assertEqual(pipeline.format_duration(0), "00:00:00")
        self.assertEqual(pipeline.format_duration(3661), "01:01:01")

    def test_progress_line_includes_resume_percentage_speed_and_eta(self) -> None:
        self.assertEqual(
            pipeline.make_progress_line(
                total_selected=100,
                resumed_success=20,
                pending_total=80,
                completed=20,
                failed=2,
                elapsed_seconds=40,
            ),
            "PROGRESS overall=40/100 percent=40.0% current=20/80 "
            "new_successes=18 new_failures=2 rate=0.500 rows/s "
            "elapsed=00:00:40 eta=00:02:00",
        )

    def test_output_layout_migrates_legacy_files_without_changing_content(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            legacy_files = {
                "annotations.csv": b"annotation-data",
                "tasks.sqlite": b"database-data",
                "attempts.jsonl": b"attempt-data",
            }
            for name, content in legacy_files.items():
                (output_dir / name).write_bytes(content)
            layout = pipeline.prepare_output_layout(output_dir)
            expected = {
                "annotations.csv": layout.results / "annotations.csv",
                "tasks.sqlite": layout.state / "tasks.sqlite",
                "attempts.jsonl": layout.audit / "attempts.jsonl",
            }
            for name, destination in expected.items():
                self.assertFalse((output_dir / name).exists())
                self.assertEqual(destination.read_bytes(), legacy_files[name])
            for directory_path in (
                layout.results,
                layout.final,
                layout.state,
                layout.audit,
                layout.logs,
            ):
                self.assertTrue(directory_path.is_dir())

    def test_all_english_contract_constants_are_exact(self) -> None:
        self.assertEqual(pipeline.SCHEMA_VERSION, "facebook-thread-annotation-v4-three-fields")
        self.assertEqual(
            pipeline.VALIDATOR_VERSION, "facebook-structural-validator-v5-three-fields"
        )
        self.assertEqual(
            pipeline.MODEL_PROTOCOL_VERSION, "facebook-thread-model-protocol-v5-three-fields"
        )
        self.assertEqual(pipeline.DEFAULT_SHEET, "comprehensive_cleaned_threads")
        self.assertEqual(
            pipeline.INPUT_HEADERS,
            [
                "topic",
                "retrieval_keyword",
                "post_text",
                "post_url",
                "post_stance_label",
                "post_stance_target",
                "post_mid",
                "thread_index",
                "thread_id",
                "participants",
                "commenter_ids",
                "commenter_usernames",
                "conversation_text",
                "root_comment_url",
                "message_count",
                "reply_count",
            ],
        )
        self.assertEqual(
            pipeline.OUTPUT_HEADERS,
            [
                "topic_relevance",
                "training_grade",
                "annotation_reason",
            ],
        )

    def test_runner_and_tests_contain_no_han_characters(self) -> None:
        han = re.compile(r"[\u3400-\u9fff]")
        for path in (PACKAGE_ROOT / "annotate_facebook_threads.py", Path(__file__)):
            self.assertIsNone(han.search(path.read_text(encoding="utf-8")), path)

    def test_task_key_changes_with_prompt_schema_model_or_input(self) -> None:
        base = pipeline.make_task_key("a\x1fb", "input", "prompt", "schema", "model")
        variants = {
            pipeline.make_task_key("a\x1fb", "input-2", "prompt", "schema", "model"),
            pipeline.make_task_key("a\x1fb", "input", "prompt-2", "schema", "model"),
            pipeline.make_task_key("a\x1fb", "input", "prompt", "schema-2", "model"),
            pipeline.make_task_key("a\x1fb", "input", "prompt", "schema", "model-2"),
        }
        self.assertEqual(len(variants), 4)
        self.assertNotIn(base, variants)
        self.assertEqual(
            base,
            pipeline.make_task_key("a\x1fb", "input", "prompt", "schema", "model"),
        )

    def test_model_protocol_hash_covers_user_and_retry_templates(self) -> None:
        base = pipeline.make_model_protocol_hash("system prompt")
        with mock.patch.object(pipeline, "USER_MESSAGE_INTRO", "changed user wrapper"):
            changed_user = pipeline.make_model_protocol_hash("system prompt")
        with mock.patch.object(pipeline, "RETRY_MESSAGE_INTRO", "changed retry wrapper"):
            changed_retry = pipeline.make_model_protocol_hash("system prompt")
        self.assertNotEqual(base, changed_user)
        self.assertNotEqual(base, changed_retry)

    def test_prompt_uses_all_english_output_contract(self) -> None:
        prompt = PROMPT_FILE.read_text(encoding="utf-8")
        self.assertIn("# Facebook Comment Thread Relevance Annotation Prompt", prompt)
        self.assertIn("Return only one three-field JSON object", prompt)
        self.assertIn("stance-expression", prompt)
        self.assertIn("one root comment plus every reply nested under that root", prompt)
        self.assertIn("Different root comments beneath the same post are separate rows", prompt)
        self.assertLessEqual(len(prompt), 12000)
        for field in ("topic", "post_text", "conversation_text"):
            self.assertIn(f"`{field}`", prompt)
        schema, _ = pipeline.load_schema(SCHEMA_PATH)
        for field, definition in schema["properties"].items():
            self.assertIn(field, prompt)
            for label in definition.get("enum", []):
                self.assertIn(label, prompt)
            if "const" in definition:
                self.assertIn(definition["const"], prompt)

    def test_launcher_reuses_post_model_location_and_vllm_method(self) -> None:
        launcher = START_SCRIPT.read_text(encoding="utf-8")
        self.assertIn(
            'MODEL_PATH="${MODEL_PATH:-${PACKAGE_ROOT}/../../models/Qwen3-32B}"',
            launcher,
        )
        self.assertIn("-m vllm.entrypoints.openai.api_server", launcher)
        self.assertIn('--served-model-name "${API_MODEL_NAME}"', launcher)
        self.assertIn("--tensor-parallel-size 4", launcher)
        self.assertIn("--dtype auto", launcher)
        self.assertIn('MAX_MODEL_LEN="${MAX_MODEL_LEN:-32768}"', launcher)
        self.assertIn('start_instance "0,1,2,3" 8000', launcher)
        self.assertIn('start_instance "4,5,6,7" 8001', launcher)

    def test_64_row_sample_is_deterministic_and_covers_required_strata(self) -> None:
        excel_rows = list(range(2, 82)) + [92, 493, 1805]
        rows = [
            make_row(excel_row, topic=f"topic-{index % 23:02d}")
            for index, excel_row in enumerate(excel_rows)
        ]
        first = pipeline.select_sample(rows, 64, 20260720)
        second = pipeline.select_sample(rows, 64, 20260720)
        self.assertEqual([row.task_key for row in first], [row.task_key for row in second])
        self.assertEqual(len(first), 64)
        self.assertEqual([row.excel_row for row in first], sorted(row.excel_row for row in first))
        self.assertTrue({11, 92, 493, 1805}.issubset({row.excel_row for row in first}))
        self.assertEqual(len({row.topic for row in first}), 23)


class VLLMClientTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.schema, _ = pipeline.load_schema(SCHEMA_PATH)

    def test_single_success(self) -> None:
        with fake_vllm([valid_related_annotation()]) as server:
            outcome = pipeline.process_row(
                make_row(), "system prompt", self.schema, make_client(server, self.schema), 0
            )
        self.assertEqual(outcome.status, "succeeded")
        self.assertEqual(len(outcome.attempts), 1)
        self.assertEqual(outcome.attempts[0].attempt_no, 1)
        self.assertEqual(len(server.requests), 1)
        request = server.requests[0]
        self.assertIn("Review the following single Facebook comment thread", request["messages"][-1]["content"])
        schema_properties = request["response_format"]["json_schema"]["schema"]["properties"]
        self.assertIn("topic_relevance", schema_properties)
        self.assertIn("training_grade", schema_properties)

    def test_preflight_accepts_expected_served_model(self) -> None:
        with fake_vllm([]) as server:
            make_client(server, self.schema).preflight([server.base_url])

    def test_preflight_rejects_a_different_served_model(self) -> None:
        with fake_vllm([], model_id="../models/Qwen3-32B") as server:
            with self.assertRaises(pipeline.PipelineError) as caught:
                make_client(server, self.schema).preflight([server.base_url])
        self.assertEqual(caught.exception.code, "preflight_failed")
        self.assertIn("Qwen3-32B", str(caught.exception))
        self.assertIn("../models/Qwen3-32B", str(caught.exception))

    def test_semantic_retry_includes_feedback_and_offsets_attempt_number(self) -> None:
        invalid = valid_related_annotation()
        invalid["topic_relevance"] = "off_topic"
        with fake_vllm([invalid, valid_related_annotation()]) as server:
            outcome = pipeline.process_row(
                make_row(),
                "system prompt",
                self.schema,
                make_client(server, self.schema),
                1,
                prior_attempt_count=7,
            )
        self.assertEqual(outcome.status, "succeeded")
        self.assertEqual([record.attempt_no for record in outcome.attempts], [8, 9])
        self.assertEqual(len(server.requests), 2)
        retry_message = server.requests[1]["messages"][-1]["content"]
        self.assertIn("The previous output failed local hard validation", retry_message)
        self.assertIn("cross.off_topic_grade", retry_message)
        self.assertIn("Q=`off_topic`", retry_message)

    def test_thinking_kwarg_fallback_does_not_retry_unrelated_http_400(self) -> None:
        planned = [
            (400, '{"error":"chat_template_kwargs unsupported"}'),
            (400, '{"error":"invalid guided JSON schema"}'),
        ]
        with fake_vllm(planned) as server, mock.patch.object(pipeline.time, "sleep", return_value=None):
            client = make_client(server, self.schema, network_retries=2)
            with self.assertRaises(pipeline.PipelineError) as caught:
                client.complete(
                    [{"role": "system", "content": "x"}],
                    semantic_round=1,
                    starting_attempt_no=1,
                )
        self.assertEqual(caught.exception.code, "http_400")
        self.assertEqual(len(server.requests), 2)


class DatabaseAndExportTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.schema, _ = pipeline.load_schema(SCHEMA_PATH)

    @staticmethod
    def attempt(attempt_no: int, prompt_tokens: int, completion_tokens: int) -> pipeline.AttemptRecord:
        return pipeline.AttemptRecord(
            attempt_no=attempt_no,
            semantic_round=1,
            endpoint="http://127.0.0.1:8000/v1",
            started_at="2026-07-20T00:00:00+00:00",
            finished_at="2026-07-20T00:00:01+00:00",
            latency_ms=1000,
            http_status=200,
            error_code=None,
            error_message=None,
            response_sha256=f"response-{attempt_no}",
            validator_errors=[] if attempt_no == 2 else ["invalid"],
            prompt_tokens=prompt_tokens,
            completion_tokens=completion_tokens,
        )

    def test_error_export_does_not_crash_on_status_field(self) -> None:
        row = make_row()
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            layout = pipeline.prepare_output_layout(output_dir)
            connection = pipeline.connect_database(layout.state / "tasks.sqlite")
            pipeline.initialize_tasks(connection, [row], "p", "ep", "s", "Qwen3-32B", False)
            pipeline.record_outcome(
                connection,
                pipeline.TaskOutcome(
                    row=row,
                    status="terminal_error",
                    annotation=None,
                    error_code="semantic_validation_failed",
                    error_message="invalid",
                    endpoint=None,
                    prompt_tokens=0,
                    completion_tokens=0,
                    attempts=[],
                    warnings=[],
                ),
            )
            _, errors, summary = pipeline.export_run_artifacts(
                connection, [row], layout, "Qwen3-32B", "2026-07-20T00:00:00+00:00"
            )
            connection.close()
            self.assertEqual(len(errors), 1)
            self.assertEqual(summary["unresolved_errors"], 1)
            with (layout.results / "errors.csv").open(
                encoding="utf-8-sig", newline=""
            ) as handle:
                exported = list(csv.DictReader(handle))
            self.assertEqual(exported[0]["task_key"], row.task_key)

    def test_resume_keeps_attempt_history_and_reports_cumulative_tokens(self) -> None:
        row = make_row()
        with tempfile.TemporaryDirectory() as directory:
            output_dir = Path(directory)
            layout = pipeline.prepare_output_layout(output_dir)
            connection = pipeline.connect_database(layout.state / "tasks.sqlite")
            pipeline.initialize_tasks(connection, [row], "p", "ep", "s", "Qwen3-32B", False)
            pipeline.record_outcome(
                connection,
                pipeline.TaskOutcome(
                    row=row,
                    status="terminal_error",
                    annotation=None,
                    error_code="semantic_validation_failed",
                    error_message="invalid",
                    endpoint="http://127.0.0.1:8000/v1",
                    prompt_tokens=10,
                    completion_tokens=5,
                    attempts=[self.attempt(1, 10, 5)],
                    warnings=[],
                ),
            )
            pipeline.initialize_tasks(connection, [row], "p", "ep", "s", "Qwen3-32B", True)
            pipeline.record_outcome(
                connection,
                pipeline.TaskOutcome(
                    row=row,
                    status="succeeded",
                    annotation=valid_related_annotation(),
                    error_code=None,
                    error_message=None,
                    endpoint="http://127.0.0.1:8000/v1",
                    prompt_tokens=20,
                    completion_tokens=7,
                    attempts=[self.attempt(2, 20, 7)],
                    warnings=[],
                ),
            )
            attempts = list(
                connection.execute(
                    "SELECT attempt_no FROM attempts WHERE task_key=? ORDER BY attempt_no",
                    (row.task_key,),
                )
            )
            _, _, summary = pipeline.export_run_artifacts(
                connection, [row], layout, "Qwen3-32B", "2026-07-20T00:00:00+00:00"
            )
            connection.close()
        self.assertEqual(attempts, [(1,), (2,)])
        self.assertEqual(summary["prompt_tokens"], 30)
        self.assertEqual(summary["completion_tokens"], 12)

    def test_combined_and_crawl_csvs_use_current_formats(self) -> None:
        rows = [make_row(2), make_row(3)]
        rows[1].values["commenter_ids"] = "user-3 | user-4"
        rows[1].values["commenter_usernames"] = "Bob | Carol"
        results = {
            2: valid_related_annotation(),
            3: valid_unrelated_annotation(),
        }
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            combined_path = root / "annotations.csv"
            crawl_path = root / "facebook_selected_commenters_for_crawl.csv"
            combined_stats = pipeline.write_combined_annotations_csv(
                rows, results, combined_path
            )
            crawl_stats = pipeline.write_crawl_input_csv(rows, results, crawl_path)
            with combined_path.open(encoding="utf-8-sig", newline="") as handle:
                combined_rows = list(csv.DictReader(handle))
            with crawl_path.open(encoding="utf-8", newline="") as handle:
                crawl_rows = list(csv.DictReader(handle))
            self.assertEqual(combined_stats["rows"], 2)
            self.assertEqual(len(combined_rows), 2)
            self.assertIn("conversation_text", combined_rows[0])
            self.assertEqual(
                list(combined_rows[0])[-3:], pipeline.OUTPUT_HEADERS
            )
            self.assertEqual(crawl_stats["selected_thread_rows"], 1)
            self.assertEqual(crawl_stats["exported_relations"], 1)
            self.assertEqual(crawl_rows[0]["commenter_id"], "user-2")
            self.assertEqual(crawl_rows[0]["source_excel_row"], "2")
            self.assertFalse(crawl_path.read_bytes().startswith(b"\xef\xbb\xbf"))


class WorkbookIntegrityTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.schema, cls.schema_hash = pipeline.load_schema(SCHEMA_PATH)
        cls.rows, cls.loader_ap_hash, cls.stats = pipeline.load_input_rows(
            INPUT_XLSX,
            pipeline.DEFAULT_SHEET,
            "effective-prompt-hash",
            cls.schema_hash,
            "Qwen3-32B",
            3505,
            40000,
            8000,
        )
        cls.source_ap_hash, cls.source_rows, cls.source_columns = pipeline.workbook_ap_hash(
            INPUT_XLSX, pipeline.DEFAULT_SHEET
        )

    def test_loader_hash_matches_writer_verification_hash(self) -> None:
        self.assertEqual(self.loader_ap_hash, self.source_ap_hash)
        self.assertEqual(self.source_rows, 3506)
        self.assertEqual(self.source_columns, 16)

    def test_each_row_is_one_root_comment_reply_chain(self) -> None:
        self.assertEqual(len(self.rows), 3505)
        self.assertEqual(len({row.values["post_mid"] for row in self.rows}), 62)
        for row in self.rows:
            conversation = row.values["conversation_text"]
            message_count = int(row.values["message_count"])
            reply_count = int(row.values["reply_count"])
            self.assertEqual(conversation.count("[ROOT COMMENT]:"), 1, row.excel_row)
            self.assertEqual(conversation.count("[REPLY]:"), reply_count, row.excel_row)
            self.assertEqual(message_count, reply_count + 1, row.excel_row)

    def test_legacy_errors_and_baseline_exactly_cover_the_frozen_workbook(self) -> None:
        selected = pipeline.select_rows_from_errors_csv(self.rows, LEGACY_ERRORS_CSV)
        baseline, rejected = pipeline.load_baseline_annotations(
            LEGACY_ANNOTATIONS_CSV, self.rows, self.schema
        )
        selected_rows = {row.excel_row for row in selected} | set(rejected)
        all_rows = {row.excel_row for row in self.rows}
        self.assertEqual(len(selected), 202)
        self.assertEqual(len(rejected), 20)
        self.assertEqual(len(selected_rows), 222)
        self.assertEqual(len(baseline), 3283)
        self.assertFalse(selected_rows & set(baseline))
        self.assertEqual(selected_rows | set(baseline), all_rows)
        self.assertNotIn(
            "relevant_without_stance",
            {value["training_grade"] for value in baseline.values()},
        )

    def test_writer_preserves_source_and_ap_and_writes_only_output_columns(self) -> None:
        source_sha_before = pipeline.sha256_file(INPUT_XLSX)
        row = self.rows[0]
        annotation = valid_unrelated_annotation()
        with tempfile.TemporaryDirectory() as directory:
            output = Path(directory) / "annotated.xlsx"
            result = pipeline.write_annotated_workbook(
                INPUT_XLSX,
                output,
                pipeline.DEFAULT_SHEET,
                self.source_ap_hash,
                self.source_rows,
                [row],
                {row.excel_row: annotation},
            )
            output_ap_hash, output_rows, output_columns = pipeline.workbook_ap_hash(
                output, pipeline.DEFAULT_SHEET
            )
            self.assertEqual(output_ap_hash, self.source_ap_hash)
            self.assertEqual(output_rows, self.source_rows)
            self.assertEqual(output_columns, 19)
            self.assertEqual(result["annotated_rows"], 1)
            from openpyxl import load_workbook

            workbook = load_workbook(output, read_only=True, data_only=False)
            sheet = workbook[pipeline.DEFAULT_SHEET]
            actual = [sheet.cell(row.excel_row, col).value for col in range(17, 20)]
            workbook.close()
            expected_excel_values = [
                None if annotation[field] == "" else annotation[field]
                for field in pipeline.OUTPUT_HEADERS
            ]
            self.assertEqual(actual, expected_excel_values)
        self.assertEqual(pipeline.sha256_file(INPUT_XLSX), source_sha_before)


class MainIntegrationTests(unittest.TestCase):
    def test_blocked_run_still_writes_combined_review_with_blank_annotation(self) -> None:
        invalid = valid_related_annotation()
        invalid["stance_expression"] = "explicit_stance"
        with tempfile.TemporaryDirectory() as directory, fake_vllm([invalid]) as server:
            output_dir = Path(directory) / "run"
            result = pipeline.main(
                [
                    "--input-xlsx",
                    str(INPUT_XLSX),
                    "--prompt-file",
                    str(PROMPT_FILE),
                    "--schema-file",
                    str(SCHEMA_PATH),
                    "--skip-manifest-check",
                    "--output-dir",
                    str(output_dir),
                    "--model",
                    "Qwen3-32B",
                    "--base-urls",
                    server.base_url,
                    "--workers",
                    "1",
                    "--network-retries",
                    "0",
                    "--semantic-retries",
                    "0",
                    "--sample-size",
                    "1",
                    "--skip-preflight",
                    "--fail-on-errors",
                ]
            )
            self.assertEqual(result, 2)
            review = output_dir / "results" / "facebook_comments_annotation_review.xlsx"
            self.assertTrue(review.is_file())
            self.assertFalse((output_dir / "final" / "facebook_comments_comprehensive_annotated.xlsx").exists())
            from openpyxl import load_workbook

            workbook = load_workbook(review, read_only=True, data_only=False)
            sheet = workbook[pipeline.DEFAULT_SHEET]
            self.assertEqual(
                [sheet.cell(1, column).value for column in range(17, 20)],
                pipeline.OUTPUT_HEADERS,
            )
            self.assertTrue(
                all(
                    cell.value is None
                    for cells in sheet.iter_rows(
                        min_row=2, max_row=sheet.max_row, min_col=17, max_col=19
                    )
                    for cell in cells
                )
            )
            workbook.close()
            with (output_dir / "results" / "errors.csv").open(
                encoding="utf-8-sig", newline=""
            ) as handle:
                error_row = next(csv.DictReader(handle))
            self.assertIn("conversation_text", error_row)

    def test_main_single_row_with_fake_vllm_writes_verified_workbook(self) -> None:
        source_sha_before = pipeline.sha256_file(INPUT_XLSX)
        with tempfile.TemporaryDirectory() as directory, fake_vllm(
            [valid_unrelated_annotation()]
        ) as server:
            output_dir = Path(directory) / "run"
            output_xlsx = output_dir / "final" / "one-row.xlsx"
            stdout = io.StringIO()
            with contextlib.redirect_stdout(stdout):
                result = pipeline.main(
                    [
                        "--input-xlsx",
                        str(INPUT_XLSX),
                        "--prompt-file",
                        str(PROMPT_FILE),
                        "--schema-file",
                        str(SCHEMA_PATH),
                        "--skip-manifest-check",
                        "--output-dir",
                        str(output_dir),
                        "--output-xlsx",
                        str(output_xlsx),
                        "--model",
                        "Qwen3-32B",
                        "--base-urls",
                        server.base_url,
                        "--workers",
                        "1",
                        "--network-retries",
                        "0",
                        "--semantic-retries",
                        "0",
                        "--sample-size",
                        "1",
                        "--skip-preflight",
                        "--fail-on-errors",
                    ]
                )
            self.assertEqual(result, 0)
            self.assertTrue(output_xlsx.is_file())
            progress_output = stdout.getvalue()
            self.assertIn(
                "RUN selected=1 resumed_success=0 pending=1 workers=1 "
                "overall=0/1 percent=0.0%",
                progress_output,
            )
            self.assertIn("PROGRESS overall=1/1 percent=100.0%", progress_output)
            self.assertIn("new_successes=1 new_failures=0", progress_output)
            self.assertIn("rate=", progress_output)
            self.assertIn("rows/s elapsed=", progress_output)
            self.assertIn("eta=00:00:00", progress_output)
            self.assertIn(
                "DONE succeeded=1 combined=1 errors=0 percent=100.0%",
                progress_output,
            )
            summary = json.loads(
                (output_dir / "results" / "run_summary.json").read_text(encoding="utf-8")
            )
            self.assertEqual(summary["succeeded"], 1)
            self.assertEqual(summary["unresolved_errors"], 0)
            self.assertTrue(summary["source_workbook_unchanged"])
            self.assertEqual(summary["processing_progress"]["selected"], 1)
            self.assertEqual(summary["processing_progress"]["resumed_success"], 0)
            self.assertEqual(summary["processing_progress"]["processed_this_run"], 1)
            self.assertEqual(summary["processing_progress"]["new_failures"], 0)
            self.assertEqual(summary["processing_progress"]["percent"], 100.0)
            self.assertGreaterEqual(summary["processing_progress"]["elapsed_seconds"], 0)
            self.assertGreaterEqual(
                summary["processing_progress"]["average_rows_per_second"], 0
            )
            self.assertTrue((output_dir / "results" / "annotations.csv").is_file())
            self.assertTrue(
                (
                    output_dir
                    / "results"
                    / "facebook_selected_commenters_for_crawl.csv"
                ).is_file()
            )
            self.assertTrue(
                (output_dir / "results" / "facebook_comments_annotation_review.xlsx").is_file()
            )
            with (output_dir / "results" / "annotations.csv").open(
                encoding="utf-8-sig", newline=""
            ) as handle:
                combined_row = next(csv.DictReader(handle))
            self.assertIn("conversation_text", combined_row)
            self.assertIn("topic_relevance", combined_row)
            self.assertNotIn("is_usable", combined_row)
            self.assertTrue((output_dir / "audit" / "attempts.jsonl").is_file())
            self.assertTrue((output_dir / "state" / "tasks.sqlite").is_file())
            output_index = (output_dir / "README.md").read_text(encoding="utf-8")
            self.assertIn("Status: **COMPLETE**", output_index)
            self.assertIn("Successful annotations: 1", output_index)
        self.assertEqual(pipeline.sha256_file(INPUT_XLSX), source_sha_before)


if __name__ == "__main__":
    unittest.main(verbosity=2)
