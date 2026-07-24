#!/usr/bin/env python3
"""Generate an audit-ready HTML report for Facebook comment annotations.

The report treats the latest three-field review workbook as the source of truth.
It never modifies the workbook or any inference output.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import random
import statistics
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


PACKAGE_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RUN_ROOT = PACKAGE_ROOT / "output" / "qwen32b_8gpu"
DEFAULT_WORKBOOK = (
    DEFAULT_RUN_ROOT
    / "results"
    / "facebook_comments_annotation_review_three_fields.xlsx"
)
DEFAULT_ERRORS = DEFAULT_RUN_ROOT / "results" / "errors.csv"
DEFAULT_CRAWL = (
    DEFAULT_RUN_ROOT / "results" / "facebook_selected_commenters_for_crawl.csv"
)
DEFAULT_RUN_SUMMARY = DEFAULT_RUN_ROOT / "results" / "run_summary.json"
DEFAULT_OUTPUT = (
    DEFAULT_RUN_ROOT / "results" / "facebook_comment_summary_report.html"
)
DEFAULT_SHEET = "comprehensive_cleaned_threads"
DEFAULT_SEED = 20260721

REQUIRED_HEADERS = {
    "topic",
    "post_text",
    "post_url",
    "post_mid",
    "thread_id",
    "conversation_text",
    "root_comment_url",
    "message_count",
    "reply_count",
    "topic_relevance",
    "training_grade",
    "annotation_reason",
}

RELEVANCE_ORDER = ["strongly_relevant", "relevant", "off_topic", "unresolved"]
GRADE_ORDER = [
    "core_usable",
    "generally_usable",
    "relevant_context_only",
    "borderline_sample",
    "unusable",
    "unresolved",
]

LABEL_ZH = {
    "strongly_relevant": "高度相关",
    "relevant": "相关",
    "off_topic": "不相关",
    "core_usable": "核心可用",
    "generally_usable": "一般可用",
    "relevant_context_only": "仅作相关语境",
    "borderline_sample": "边界样本",
    "unusable": "不可用",
    "unresolved": "未解决",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Generate the Facebook comment annotation HTML report."
    )
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--errors-csv", type=Path, default=DEFAULT_ERRORS)
    parser.add_argument("--crawl-csv", type=Path, default=DEFAULT_CRAWL)
    parser.add_argument("--run-summary", type=Path, default=DEFAULT_RUN_SUMMARY)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument(
        "--sample-translations",
        type=Path,
        help=(
            "Optional JSON overlay keyed by thread_id. When omitted, the report "
            "looks for sample_translations_zh.json beside the output HTML."
        ),
    )
    parser.add_argument("--seed", type=int, default=DEFAULT_SEED)
    parser.add_argument("--usable-samples", type=int, default=12)
    parser.add_argument("--unusable-samples", type=int, default=8)
    parser.add_argument("--unresolved-samples", type=int, default=4)
    return parser.parse_args()


def esc(value: Any) -> str:
    return html.escape("" if value is None else str(value), quote=True)


def as_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def as_int(value: Any) -> int:
    try:
        return int(value or 0)
    except (TypeError, ValueError):
        return 0


def fmt_number(value: int | float) -> str:
    return f"{value:,.0f}"


def pct(numerator: int, denominator: int) -> str:
    return "-" if denominator <= 0 else f"{numerator / denominator:.1%}"


def quantile(values: list[int], q: float) -> int:
    if not values:
        return 0
    ordered = sorted(values)
    index = round((len(ordered) - 1) * q)
    return int(ordered[index])


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_json(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {}
    value = json.loads(path.read_text(encoding="utf-8-sig"))
    return value if isinstance(value, dict) else {}


def read_sample_translations(path: Path) -> dict[str, dict[str, str]]:
    payload = read_json(path)
    translations = payload.get("translations", payload)
    if not isinstance(translations, dict):
        raise SystemExit(f"Invalid sample translation overlay: {path}")
    cleaned: dict[str, dict[str, str]] = {}
    for thread_id, value in translations.items():
        if thread_id == "_meta" or not isinstance(value, dict):
            continue
        cleaned[str(thread_id)] = {
            key: as_text(value.get(key))
            for key in ("post_text_zh", "conversation_text_zh", "annotation_reason_zh")
        }
    return cleaned


def read_topic_translations(path: Path) -> dict[str, str]:
    payload = read_json(path)
    translations = payload.get("topic_translations", {})
    if not isinstance(translations, dict):
        raise SystemExit(f"Invalid topic translation overlay: {path}")
    return {
        as_text(topic): as_text(translation)
        for topic, translation in translations.items()
        if as_text(topic) and as_text(translation)
    }


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_workbook(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {sheet_name}")
    sheet = workbook[sheet_name]
    headers = [as_text(sheet.cell(1, column).value) for column in range(1, sheet.max_column + 1)]
    missing = sorted(REQUIRED_HEADERS.difference(headers))
    if missing:
        raise SystemExit(f"Workbook is missing required headers: {missing}")
    rows = [
        dict(zip(headers, values))
        for values in sheet.iter_rows(min_row=2, max_col=len(headers), values_only=True)
    ]
    workbook.close()
    return rows


def is_completed(row: dict[str, Any]) -> bool:
    return bool(as_text(row.get("topic_relevance")))


def is_selected(row: dict[str, Any]) -> bool:
    return (
        as_text(row.get("topic_relevance")) in {"strongly_relevant", "relevant"}
        and as_text(row.get("training_grade")) != "unusable"
    )


def stratified_sample(
    rows: Iterable[dict[str, Any]], count: int, seed: int
) -> list[dict[str, Any]]:
    groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        groups[as_text(row.get("topic"))].append(row)
    rng = random.Random(seed)
    topics = sorted(groups)
    rng.shuffle(topics)
    for topic in topics:
        groups[topic].sort(key=lambda item: (as_text(item.get("thread_id")), as_text(item.get("post_mid"))))
        rng.shuffle(groups[topic])
    selected: list[dict[str, Any]] = []
    while len(selected) < count:
        added = False
        for topic in topics:
            if groups[topic]:
                selected.append(groups[topic].pop())
                added = True
                if len(selected) >= count:
                    break
        if not added:
            break
    return selected


def truncate(value: Any, limit: int) -> str:
    text = as_text(value)
    if len(text) <= limit:
        return text
    return text[: limit - 1].rstrip() + "…"


def safe_link(url: Any, label: str) -> str:
    value = as_text(url)
    if not value.lower().startswith(("http://", "https://")):
        return ""
    return f'<a href="{esc(value)}" target="_blank" rel="noopener noreferrer">{esc(label)}</a>'


def metric_card(label: str, value: str, note: str, tone: str = "") -> str:
    return (
        f'<div class="metric {esc(tone)}"><div class="metric-label">{esc(label)}</div>'
        f'<div class="metric-value">{esc(value)}</div><div class="metric-note">{esc(note)}</div></div>'
    )


def render_table(headers: list[str], rows: list[list[Any]], numeric: set[int] | None = None) -> str:
    numeric = numeric or set()
    head = "".join(
        f'<th class="{"num" if index in numeric else ""}">{esc(header)}</th>'
        for index, header in enumerate(headers)
    )
    body = []
    for row in rows:
        cells = "".join(
            f'<td class="{"num" if index in numeric else ""}">{value if isinstance(value, Html) else esc(value)}</td>'
            for index, value in enumerate(row)
        )
        body.append(f"<tr>{cells}</tr>")
    return f"<div class=\"table-wrap\"><table><thead><tr>{head}</tr></thead><tbody>{''.join(body)}</tbody></table></div>"


class Html(str):
    """Marker for trusted HTML fragments assembled by this script."""


def label_tag(value: Any) -> Html:
    label = as_text(value) or "unresolved"
    tone = "good" if label in {"strongly_relevant", "core_usable", "generally_usable"} else ""
    if label in {"relevant", "relevant_context_only", "borderline_sample"}:
        tone = "warn"
    if label in {"off_topic", "unusable", "unresolved"}:
        tone = "bad"
    title = LABEL_ZH.get(label, label)
    return Html(f'<span class="tag {tone}" title="{esc(label)}">{esc(title)}</span>')


def sample_card(
    row: dict[str, Any],
    index: int,
    kind: str,
    topic_number: int,
    topic_translation: str,
    translation: dict[str, str] | None = None,
) -> str:
    translation = translation or {}
    relevance = as_text(row.get("topic_relevance")) or "unresolved"
    grade = as_text(row.get("training_grade")) or "unresolved"
    links = " · ".join(
        item
        for item in (
            safe_link(row.get("post_url"), "原始帖子"),
            safe_link(row.get("root_comment_url"), "根评论"),
        )
        if item
    )
    reason = truncate(row.get("annotation_reason"), 800) or "该线程尚未得到有效标注。"
    def translated_block(label: str, key: str, extra_class: str = "") -> str:
        value = translation.get(key, "")
        if not value:
            return ""
        classes = f"sample-block {extra_class} translation".strip()
        return f'<div class="{classes}"><strong>{esc(label)}</strong>{esc(value)}</div>'

    post_translation = translated_block(
        "帖子中文（辅助理解）：", "post_text_zh"
    )
    conversation_translation = translated_block(
        "评论线程中文（辅助理解）：", "conversation_text_zh", "conversation"
    )
    reason_translation = translated_block(
        "标注理由中文（辅助理解）：", "annotation_reason_zh", "reason"
    )
    return f"""
<article class="sample {esc(kind)}">
  <div class="sample-top"><strong>样本 #{index} · Topic #{topic_number}: {esc(as_text(row.get('topic')))}<span class="topic-zh">{esc(topic_translation)}</span></strong><span>{label_tag(relevance)} {label_tag(grade)}</span></div>
  <div class="sample-meta">thread_id: {esc(as_text(row.get('thread_id')))} · post_mid: {esc(as_text(row.get('post_mid')))} · messages: {as_int(row.get('message_count'))} · replies: {as_int(row.get('reply_count'))}</div>
  <div class="sample-block"><strong>帖子：</strong>{esc(truncate(row.get('post_text'), 1000))}</div>
  {post_translation}
  <div class="sample-block conversation"><strong>评论线程：</strong>{esc(truncate(row.get('conversation_text'), 2200))}</div>
  {conversation_translation}
  <div class="sample-block reason"><strong>模型理由：</strong>{esc(reason)}</div>
  {reason_translation}
  <div class="sample-links">{links}</div>
</article>"""


def main() -> None:
    args = parse_args()
    if not args.workbook.exists():
        raise SystemExit(f"Workbook not found: {args.workbook}")

    source_hash_before = sha256_file(args.workbook)
    rows = read_workbook(args.workbook, args.sheet)
    source_hash_after = sha256_file(args.workbook)
    if source_hash_before != source_hash_after:
        raise SystemExit("Source workbook changed while generating the report")

    errors = read_csv(args.errors_csv)
    crawl_rows = read_csv(args.crawl_csv)
    run_summary = read_json(args.run_summary)
    translation_path = (
        args.sample_translations
        if args.sample_translations is not None
        else args.output.parent / "sample_translations_zh.json"
    )
    sample_translations = (
        read_sample_translations(translation_path) if translation_path.exists() else {}
    )
    topic_translations = (
        read_topic_translations(translation_path) if translation_path.exists() else {}
    )

    total = len(rows)
    completed_rows = [row for row in rows if is_completed(row)]
    unresolved_rows = [row for row in rows if not is_completed(row)]
    selected_rows = [row for row in completed_rows if is_selected(row)]
    excluded_rows = [row for row in completed_rows if not is_selected(row)]

    relevance_counts = Counter(
        as_text(row.get("topic_relevance")) or "unresolved" for row in rows
    )
    grade_counts = Counter(as_text(row.get("training_grade")) or "unresolved" for row in rows)
    topic_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        topic_groups[as_text(row.get("topic"))].append(row)

    ordered_topics = sorted(
        topic_groups.items(), key=lambda item: (-len(item[1]), item[0])
    )
    topic_numbers = {
        topic: index for index, (topic, _) in enumerate(ordered_topics, 1)
    }
    topic_rows: list[list[Any]] = []
    for topic, topic_items in ordered_topics:
        success = sum(is_completed(row) for row in topic_items)
        selected = sum(is_selected(row) for row in topic_items)
        unresolved = len(topic_items) - success
        topic_row: list[Any] = [
            topic_numbers[topic],
            Html(
                f'<span class="topic">{esc(topic)}'
                f'<span class="topic-zh">{esc(topic_translations.get(topic))}</span></span>'
            ),
            fmt_number(len(topic_items)),
            fmt_number(selected),
            pct(selected, success),
        ]
        if unresolved_rows:
            topic_row.append(fmt_number(unresolved))
        topic_rows.append(topic_row)

    message_counts = [as_int(row.get("message_count")) for row in rows]
    reply_counts = [as_int(row.get("reply_count")) for row in rows]
    unique_posts = {as_text(row.get("post_mid")) for row in rows if as_text(row.get("post_mid"))}
    unique_threads = {as_text(row.get("thread_id")) for row in rows if as_text(row.get("thread_id"))}
    multi_message_threads = sum(value > 1 for value in message_counts)

    error_counts = Counter(row.get("error_code") or "unknown" for row in errors)
    error_table_rows = [
        [code, fmt_number(count), pct(count, len(errors))]
        for code, count in error_counts.most_common()
    ] or [["无", "0", "-"]]

    relevance_labels = (
        RELEVANCE_ORDER
        if unresolved_rows
        else [label for label in RELEVANCE_ORDER if label != "unresolved"]
    )
    grade_labels = (
        GRADE_ORDER
        if unresolved_rows
        else [label for label in GRADE_ORDER if label != "unresolved"]
    )
    relevance_table_rows = [
        [label_tag(label), fmt_number(relevance_counts[label]), pct(relevance_counts[label], total)]
        for label in relevance_labels
    ]
    grade_table_rows = [
        [label_tag(label), fmt_number(grade_counts[label]), pct(grade_counts[label], total)]
        for label in grade_labels
    ]

    crawl_commenters = {
        row.get("commenter_id", "") for row in crawl_rows if row.get("commenter_id", "")
    }
    crawl_threads = {row.get("thread_id", "") for row in crawl_rows if row.get("thread_id", "")}
    crawl_posts = {row.get("post_mid", "") for row in crawl_rows if row.get("post_mid", "")}

    usable_samples = stratified_sample(selected_rows, args.usable_samples, args.seed)
    unusable_samples = stratified_sample(excluded_rows, args.unusable_samples, args.seed + 1)
    unresolved_samples = stratified_sample(
        unresolved_rows, args.unresolved_samples, args.seed + 2
    )

    report_samples = usable_samples + unusable_samples + unresolved_samples
    if topic_translations:
        missing_topics = [topic for topic, _ in ordered_topics if topic not in topic_translations]
        if missing_topics:
            raise SystemExit(
                "Topic translation overlay is incomplete:\n" + "\n".join(missing_topics)
            )
    if sample_translations:
        missing_translation_fields: list[str] = []
        for row in report_samples:
            thread_id = as_text(row.get("thread_id"))
            translated = sample_translations.get(thread_id, {})
            missing = [
                key
                for key in (
                    "post_text_zh",
                    "conversation_text_zh",
                    "annotation_reason_zh",
                )
                if not translated.get(key)
            ]
            if missing:
                missing_translation_fields.append(
                    f"{thread_id}: {', '.join(missing)}"
                )
        if missing_translation_fields:
            raise SystemExit(
                "Sample translation overlay is incomplete:\n"
                + "\n".join(missing_translation_fields)
            )

    sample_html = "".join(
        sample_card(
            row,
            index,
            "usable",
            topic_numbers[as_text(row.get("topic"))],
            topic_translations.get(as_text(row.get("topic")), ""),
            sample_translations.get(as_text(row.get("thread_id"))),
        )
        for index, row in enumerate(usable_samples, 1)
    )
    unusable_html = "".join(
        sample_card(
            row,
            index,
            "excluded",
            topic_numbers[as_text(row.get("topic"))],
            topic_translations.get(as_text(row.get("topic")), ""),
            sample_translations.get(as_text(row.get("thread_id"))),
        )
        for index, row in enumerate(unusable_samples, 1)
    )
    unresolved_html = "".join(
        sample_card(
            row,
            index,
            "unresolved",
            topic_numbers[as_text(row.get("topic"))],
            topic_translations.get(as_text(row.get("topic")), ""),
            sample_translations.get(as_text(row.get("thread_id"))),
        )
        for index, row in enumerate(unresolved_samples, 1)
    )

    model = as_text(run_summary.get("model")) or "Qwen3-32B"
    generated_at = datetime.now().astimezone().isoformat(timespec="seconds")
    status = "尚未最终完成" if unresolved_rows else "已完成"
    status_tone = "bad" if unresolved_rows else "good"
    unresolved_metric = (
        metric_card(
            "未解决",
            fmt_number(len(unresolved_rows)),
            f"占全部线程 {pct(len(unresolved_rows), total)}",
            status_tone,
        )
        if unresolved_rows
        else ""
    )
    metric_columns = 5 if unresolved_rows else 4
    integrity_callout = (
        f'<div class="callout bad"><strong>运行完整性：</strong>{fmt_number(len(completed_rows))} '
        f'条有效标注，{fmt_number(len(unresolved_rows))} 条仍为空；因此当前报告是复核版，不是最终封版。</div>'
        if unresolved_rows
        else f'<div class="callout good"><strong>运行完整性：</strong>全部 {fmt_number(total)} 条线程均已获得有效标注，当前报告为最终完整版本。</div>'
    )
    error_section = (
        f'<section><div class="section-head"><div><h2>未解决错误</h2><p class="hint">来自 errors.csv；这些行没有进入当前保留集合。</p></div></div>'
        f'{render_table(["错误代码", "数量", "占未解决"], error_table_rows, {1})}</section>'
        if unresolved_rows
        else ""
    )
    unresolved_sample_section = (
        f'<section><div class="section-head"><div><h2>未解决线程固定抽样</h2><p class="hint">这些线程保留原始帖子和评论内容，但标注字段为空。</p></div></div><div class="samples">{unresolved_html}</div></section>'
        if unresolved_rows
        else ""
    )
    topic_headers = ["序号", "Topic", "总线程", "保留", "成功口径保留率"]
    topic_numeric = {0, 2, 3}
    if unresolved_rows:
        topic_headers.append("未解决")
        topic_numeric.add(5)
    footer_error_source = (
        f'；错误：<code>{esc(args.errors_csv)}</code>' if unresolved_rows else ""
    )

    document = f"""<!doctype html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Facebook 评论筛选报告</title>
<style>
:root {{ --ink:#17212b; --muted:#61707d; --line:#dbe2e8; --panel:#fff; --page:#f4f7f9; --blue:#0b5f7e; --green:#1f7a45; --amber:#93630c; --red:#a33a3a; --soft-green:#e9f6ee; --soft-amber:#fff4d9; --soft-red:#fdecec; }}
* {{ box-sizing:border-box; }} body {{ margin:0; color:var(--ink); background:var(--page); font:14px/1.55 -apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif; }}
header {{ color:white; background:linear-gradient(125deg,#10384b,#0b637d 60%,#2d7c72); }} .header-inner,main {{ width:min(1180px,calc(100% - 32px)); margin:auto; }} .header-inner {{ padding:34px 0 30px; }}
h1 {{ margin:0 0 7px; font-size:30px; }} h2 {{ margin:0; font-size:20px; }} .subtitle {{ margin:0; color:#dcebf0; }}
.status {{ display:inline-flex; margin-top:16px; padding:6px 10px; border-radius:999px; background:#fff1; border:1px solid #ffffff55; font-weight:700; }}
main {{ padding:22px 0 38px; }} section {{ margin:0 0 18px; padding:18px; border:1px solid var(--line); border-radius:12px; background:var(--panel); box-shadow:0 7px 20px #1327330c; }}
.section-head {{ display:flex; justify-content:space-between; gap:12px; margin-bottom:13px; }} .hint,.metric-note,.sample-meta {{ color:var(--muted); }} .hint {{ margin:3px 0 0; font-size:13px; }}
.metrics {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:12px; }} .metric {{ padding:14px; border:1px solid var(--line); border-top:4px solid #8796a2; border-radius:9px; }} .metric.good {{ border-top-color:var(--green); }} .metric.warn {{ border-top-color:var(--amber); }} .metric.bad {{ border-top-color:var(--red); }}
.metric-label {{ color:var(--muted); font-size:12px; font-weight:700; }} .metric-value {{ margin:3px 0; font-size:27px; font-weight:750; font-variant-numeric:tabular-nums; }} .metric-note {{ font-size:12px; }}
.grid-2 {{ display:grid; grid-template-columns:1fr 1fr; gap:14px; }} .callout {{ padding:12px 13px; border-left:4px solid var(--blue); border-radius:6px; background:#eef7fa; }} .callout.good {{ border-color:var(--green); background:var(--soft-green); }} .callout.warn {{ border-color:var(--amber); background:var(--soft-amber); }} .callout.bad {{ border-color:var(--red); background:var(--soft-red); }}
.callout-list {{ display:grid; gap:9px; }} .logic-list {{ margin:0; padding-left:22px; }} .logic-list li {{ margin:7px 0; padding-left:3px; }} .table-wrap {{ overflow:auto; border:1px solid var(--line); border-radius:8px; }} table {{ width:100%; border-collapse:collapse; }} th,td {{ padding:9px 10px; border-bottom:1px solid var(--line); text-align:left; vertical-align:top; font-size:13px; }} th {{ background:#edf2f5; color:#3e4b57; white-space:nowrap; }} tr:last-child td {{ border-bottom:0; }} .num {{ text-align:right; font-variant-numeric:tabular-nums; }} .topic {{ display:block; min-width:300px; max-width:520px; }} .topic-zh {{ display:block; margin-top:3px; color:var(--muted); font-size:12px; font-weight:500; }}
.tag {{ display:inline-flex; align-items:center; min-height:22px; margin:1px 2px; padding:2px 7px; border-radius:999px; background:#edf1f4; color:#34424e; font-size:12px; font-weight:700; }} .tag.good {{ background:var(--soft-green); color:#1d6a39; }} .tag.warn {{ background:var(--soft-amber); color:#744a09; }} .tag.bad {{ background:var(--soft-red); color:#963232; }}
.samples {{ display:grid; gap:11px; }} .sample {{ padding:14px; border:1px solid var(--line); border-left:4px solid var(--green); border-radius:9px; background:#fcfdfd; }} .sample.excluded {{ border-left-color:var(--amber); }} .sample.unresolved {{ border-left-color:var(--red); }} .sample-top {{ display:flex; justify-content:space-between; gap:10px; align-items:flex-start; }} .sample-meta {{ margin:5px 0 8px; font-size:12px; overflow-wrap:anywhere; }} .sample-block {{ margin-top:7px; white-space:pre-wrap; overflow-wrap:anywhere; }} .conversation {{ padding:10px; border-radius:7px; background:#f3f6f8; }} .reason {{ color:#294b3b; }} .sample-links {{ margin-top:9px; }} a {{ color:var(--blue); }}
.translation {{ margin-top:5px; padding:9px 10px; border-left:3px solid #76a9b8; border-radius:6px; background:#eef7fa; color:#173f4c; }} .conversation.translation {{ background:#eef7fa; }} .reason.translation {{ color:#173f4c; }}
.footer {{ color:var(--muted); font-size:12px; overflow-wrap:anywhere; }} code {{ padding:1px 4px; border-radius:4px; background:#eef2f5; }}
@media (max-width:900px) {{ .metrics,.grid-2 {{ grid-template-columns:1fr 1fr; }} .sample-top {{ display:block; }} }} @media (max-width:600px) {{ .metrics,.grid-2 {{ grid-template-columns:1fr; }} h1 {{ font-size:25px; }} }}
</style>
</head>
<body>
<header><div class="header-inner">
  <h1>Facebook 评论筛选报告</h1>
  <p class="subtitle">标注单位：一个根评论及其全部回复。生成时间：{esc(generated_at)}</p>
  <div class="status">当前状态：{esc(status)}</div>
</div></header>
<main>
<section><div class="metrics" style="grid-template-columns:repeat({metric_columns},minmax(0,1fr))">
  {metric_card('评论线程总数', fmt_number(total), f'{fmt_number(len(unique_posts))} 个帖子 / {fmt_number(len(unique_threads))} 个唯一线程')}
  {metric_card('Topic 总数', fmt_number(len(topic_groups)), '下方汇总表按规模排序并连续编号')}
  {metric_card('成功标注', fmt_number(len(completed_rows)), f'完成率 {pct(len(completed_rows), total)}', 'good')}
  {metric_card('当前保留', fmt_number(len(selected_rows)), f'占成功标注 {pct(len(selected_rows), len(completed_rows))}', 'warn')}
  {unresolved_metric}
</div></section>

<section><div class="section-head"><div><h2>此次筛选逻辑与结果口径</h2><p class="hint">这是宽松的第一轮相关性筛选；只判断主题相关性和训练可用性，不标注立场或立场变化。</p></div></div>
<div class="callout-list">
  {integrity_callout}
  <div class="callout"><ol class="logic-list">
    <li><strong>输入范围：</strong>从帖子级清洗后保留的 Facebook 帖子出发，本报告覆盖 {fmt_number(len(unique_posts))} 个帖子下的评论。</li>
    <li><strong>判断单位：</strong>一个根评论及其全部回复构成一个评论线程；同一帖子下不同根评论分别判断。</li>
    <li><strong>模型输入：</strong>将指定 Topic、原始帖子和完整评论线程一起交给 {esc(model)}，输出 <code>topic_relevance</code>、<code>training_grade</code> 和 <code>annotation_reason</code> 三个字段。</li>
    <li><strong>相关性口径：</strong><code>strongly_relevant</code> 表示直接涉及主题或核心政策对象；<code>relevant</code> 表示存在明确的间接或相邻关联；<code>off_topic</code> 表示没有可靠关联或仅为噪声。</li>
    <li><strong>最终保留：</strong><code>topic_relevance ∈ {{strongly_relevant, relevant}}</code> 且 <code>training_grade != unusable</code>。本次保留 {fmt_number(len(selected_rows))} 个线程，排除 {fmt_number(len(excluded_rows))} 个线程。</li>
    <li><strong>程序校验边界：</strong>程序只校验 JSON 格式、枚举值和字段间机械一致性，不根据评论内容重新判定标签是否正确。</li>
  </ol></div>
  <div class="callout warn"><strong>人工复核：</strong><code>relevant_context_only</code> 和 <code>borderline_sample</code> 按当前宽松口径保留，应优先人工审查；立场及立场变化留待后续人工标注。</div>
  <div class="callout"><strong>模型与来源：</strong>{esc(model)}；源工作簿 SHA-256 为 <code>{esc(source_hash_before)}</code>，生成前后未改变。</div>
</div></section>

<section><div class="grid-2"><div><div class="section-head"><div><h2>主题相关性</h2><p class="hint">占全部 3,505 个线程。</p></div></div>{render_table(['标签','数量','占比'], relevance_table_rows, {1})}</div>
<div><div class="section-head"><div><h2>训练可用等级</h2><p class="hint">未解决行单列显示。</p></div></div>{render_table(['标签','数量','占比'], grade_table_rows, {1})}</div></div></section>

<section><div class="section-head"><div><h2>线程结构与后续爬取</h2><p class="hint">消息数来自工作簿；评论者统计来自当前 crawl CSV。</p></div></div>
<div class="metrics">
  {metric_card('消息总数', fmt_number(sum(message_counts)), f'回复 {fmt_number(sum(reply_counts))} 条')}
  {metric_card('含回复线程', fmt_number(multi_message_threads), f'占全部线程 {pct(multi_message_threads, total)}')}
  {metric_card('线程消息 P50 / P90 / Max', f'{quantile(message_counts, .5)} / {quantile(message_counts, .9)} / {max(message_counts) if message_counts else 0}', '按线程统计')}
  {metric_card('可继续爬取的评论者', fmt_number(len(crawl_commenters)), f'{fmt_number(len(crawl_rows))} 条关系 / {fmt_number(len(crawl_threads))} 个线程 / {fmt_number(len(crawl_posts))} 个帖子', 'good')}
</div></section>

{error_section}

<section><div class="section-head"><div><h2>按 Topic 汇总（共 {len(topic_groups)} 个）</h2><p class="hint">Topic 按线程数量从多到少编号；表格已隐藏“成功”列，保留率仍以该 Topic 的成功标注行作为分母。</p></div></div>{render_table(topic_headers, topic_rows, topic_numeric)}</section>

<section><div class="section-head"><div><h2>可用线程固定抽样</h2><p class="hint">跨 Topic 固定抽取 {len(usable_samples)} 条；随机种子 {args.seed}。中文为辅助理解译文，审计时仍以英文原文为准。</p></div></div><div class="samples">{sample_html}</div></section>
<section><div class="section-head"><div><h2>排除线程固定抽样</h2><p class="hint">从成功但未通过保留规则的线程中固定抽取 {len(unusable_samples)} 条。中文为辅助理解译文，审计时仍以英文原文为准。</p></div></div><div class="samples">{unusable_html}</div></section>
{unresolved_sample_section}

<section class="footer">报告由 <code>scripts/generate_facebook_comment_report.py</code> 生成。源数据：<code>{esc(args.workbook)}</code>{footer_error_source}；下游评论者：<code>{esc(args.crawl_csv)}</code>。</section>
</main>
</body>
</html>
"""

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(document, encoding="utf-8")
    print(f"Wrote {args.output}")
    print(
        f"Threads={total:,}; completed={len(completed_rows):,}; "
        f"selected={len(selected_rows):,}; unresolved={len(unresolved_rows):,}"
    )
    print(
        f"Samples={len(usable_samples) + len(unusable_samples) + len(unresolved_samples):,}; "
        f"source_sha256={source_hash_before}"
    )


if __name__ == "__main__":
    main()
